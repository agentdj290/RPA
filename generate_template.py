# generate_template.py
import openpyxl
from core import CommandType

def create_cmd_template(filename="cmd_template.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RPA指令"

    # 表头（必须与你的 validate_commands 兼容）
    ws.append(["操作类型", "参数", "重试次数"])

    # 所有命令类型的完整示例（按 CommandType 枚举顺序）
    examples = [
        # (操作类型值, 参数示例, 重试次数, 说明)
        (CommandType.LEFT_CLICK.value, "button_ok.png", 3, "左键单击：点击指定图片位置"),
        (CommandType.LEFT_DOUBLE_CLICK.value, "file_icon.png", 3, "左键双击：常用于打开文件/图标"),
        (CommandType.RIGHT_CLICK.value, "desktop_empty.png", 3, "右键单击：弹出上下文菜单"),
        (CommandType.TEXT_INPUT.value, "用户名或密码内容", 1, "文本输入：自动粘贴文本（支持中文）"),
        (CommandType.WAIT.value, "2.0", 1, "等待：单位为秒，可填小数如 0.5"),
        (CommandType.SCROLL.value, "-300", 1, "滚轮滑动：正数向上，负数向下（单位）"),
        (CommandType.COPY.value, "", 1, "复制：执行 Ctrl+C（无需参数）"),
        (CommandType.PASTE.value, "", 1, "粘贴：执行 Ctrl+V（无需参数）"),
        (CommandType.DELETE.value, "", 1, "删除：按下 Delete 键（无需参数）"),
        (CommandType.HOTKEY.value, "ctrl+s", 1, "快捷键：按键组合，用 '+' 连接，如 ctrl+c / alt+f4 / tab"),
        (CommandType.ACTIVATE_WINDOW.value, "微信", 2, "窗口激活：根据标题关键词激活窗口（仅 Windows）"),
    ]

    for cmd_type, param, retry, comment in examples:
        ws.append([cmd_type, param, retry])
        ws.cell(row=ws.max_row, column=4).value = f"← {comment}"

    # 调整列宽提升可读性
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 40

    wb.save(filename)
    print(f"✅ 完整模板已生成: {filename}")
    print("💡 提示：")
    print("   - 图片需与 Excel 文件放在同一目录，或使用相对路径（如 images/btn.png）")
    print("   - 快捷键请用小写，多个键用 '+' 连接")
    print("   - 空参数的命令（复制/粘贴/删除）请留空，不要填任何内容")

if __name__ == "__main__":
    create_cmd_template()
